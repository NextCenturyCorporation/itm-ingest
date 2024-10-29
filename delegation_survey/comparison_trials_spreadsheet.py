import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime

def calculate_review_time(interactions, dm_name):
    total_time = 0
    open_timestamp = None
    
    # sort interactions by timestamp to ensure proper pairing
    sorted_interactions = sorted(interactions, key=lambda x: x['timestamp'])
    
    for interaction in sorted_interactions:
        if interaction['dmName'] != dm_name:
            continue
            
        current_time = datetime.strptime(interaction['timestamp'], '%Y-%m-%dT%H:%M:%S.%fZ')
        
        if interaction['actionName'] == 'Opened DM Review Modal':
            open_timestamp = current_time
        elif interaction['actionName'] == 'Closed DM Review Modal' and open_timestamp is not None:
            time_diff = (current_time - open_timestamp).total_seconds()
            total_time += time_diff
            open_timestamp = None
    
    return round(total_time, 2)

def comparison_trials_spreadsheet(mongo_db, output_file='comparison_trials.xlsx'):
    # read the RQ data file
    rq_data = pd.read_excel('RQ-1_and_RQ-3_data.xlsx')
    
    # Convert relevant columns to lowercase for case-insensitive matching
    rq_data['Delegator_ID'] = rq_data['Delegator_ID'].astype(str)
    rq_data['TA1_Name'] = rq_data['TA1_Name'].str.upper()
    rq_data['TA2_Name'] = rq_data['TA2_Name'].str.lower()
    rq_data['ADM_Aligned_Status (Baseline/Misaligned/Aligned)'] = rq_data['ADM_Aligned_Status (Baseline/Misaligned/Aligned)'].str.lower()
    
    survey_results = mongo_db['surveyResults']
    eval_4_results = list(survey_results.find({"results.evalNumber": 4}))

    keys_to_remove = {
        'VR Page', 'user', 'PID Warning', 'Participant ID Page', 'browserInfo', 
        'Survey Introduction', 'Note page', 'Post-Scenario Measures'
    }

    table_data = []
    processed_participants = set()

    for entry in eval_4_results:
        results = entry['results']
        try:
            pid = results['Participant ID Page']['questions']['Participant ID']['response']
            if 'test' in pid.lower():
                continue
            processed_participants.add(pid)
        except KeyError:
            continue
        
        results = {k: v for k, v in results.items() if isinstance(v, dict) and k not in keys_to_remove}
        
        scenario_groups = defaultdict(list)
        for k, v in results.items():
            if 'scenarioIndex' in v:
                scenario_index = v['scenarioIndex']
                scenario_groups[scenario_index].append((k, v))

        for block_num, (scenario_index, group_data) in enumerate(scenario_groups.items(), 1):
            ta1 = 'ADEPT' if 'DryRun' in scenario_index else 'ST'
            ta2 = 'Parallax' if group_data[0][1].get('admAuthor', '') == 'TAD' else group_data[0][1].get('admAuthor', '')
            
            row_data = {
                'Participant ID': pid,
                'Block': f'Block {block_num}',
                'TA1': ta1,
                'TA2': ta2,
                'Scenario': scenario_index
            }

            dm_names = {}

            dm_count = len(group_data) - 1 if len(group_data) > 2 else len(group_data)
            for i in range(3):
                if i < dm_count:
                    entry = group_data[i][1]
                    dm_name = entry.get('pageName', '')
                    dm_type = entry.get('admAlignment', '').lower()  # Convert to lowercase
                    row_data[f'DM{i+1}'] = dm_name
                    row_data[f'DM{i+1}_Type'] = dm_type
                    row_data[f'DM{i+1}_Time'] = entry.get('timeSpentOnPage', '')
                    row_data[f'DM{i+1}_View'] = 0
                    row_data[f'DM{i+1}_Review_Time'] = 0
                    row_data[f'DM{i+1}_Alignment_Score'] = ''
                    dm_names[dm_name] = i+1

                    matching_rows = rq_data[
                        (rq_data['Delegator_ID'] == str(pid)) &
                        (rq_data['TA1_Name'].str.upper() == ta1.upper()) &
                        (rq_data['TA2_Name'].str.lower() == ta2.lower()) &
                        (rq_data['ADM_Aligned_Status (Baseline/Misaligned/Aligned)'].str.lower() == dm_type.lower())
                    ]
                    
                    if not matching_rows.empty:
                        alignment_score = matching_rows['Alignment score (Delegator|Observed_ADM (target))'].iloc[0]
                        row_data[f'DM{i+1}_Alignment_Score'] = alignment_score
                else:
                    row_data[f'DM{i+1}'] = ''
                    row_data[f'DM{i+1}_Type'] = ''
                    row_data[f'DM{i+1}_Time'] = ''
                    row_data[f'DM{i+1}_View'] = 0
                    row_data[f'DM{i+1}_Review_Time'] = 0
                    row_data[f'DM{i+1}_Alignment_Score'] = ''

            if len(group_data) > 2:
                compare_entry = group_data[-1][1]
                row_data['Compare_Time'] = compare_entry.get('timeSpentOnPage', '')

                questions = compare_entry.get('questions', {})
                review_key = next((key for key in questions.keys() if 'Review' in key), None)

                if review_key:
                    interactions = questions[review_key].get('response', [])
                    
                    for dm_name, dm_number in dm_names.items():
                        was_viewed = any(
                            interaction['dmName'] == dm_name and 
                            interaction['actionName'] == 'Opened DM Review Modal'
                            for interaction in interactions
                        )
                        row_data[f'DM{dm_number}_View'] = 1 if was_viewed else 0
                        
                        total_review_time = calculate_review_time(interactions, dm_name)
                        row_data[f'DM{dm_number}_Review_Time'] = total_review_time
            else:
                row_data['Compare_Time'] = ''

            table_data.append(row_data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Trials"

    headers = ['Participant ID', 'Block', 'TA1', 'TA2', 'Scenario',
               'DM1', 'DM1_Type', 'DM1_Alignment_Score', 'DM1_Time', 'DM1_View', 'DM1_Review_Time',
               'DM2', 'DM2_Type', 'DM2_Alignment_Score', 'DM2_Time', 'DM2_View', 'DM2_Review_Time',
               'DM3', 'DM3_Type', 'DM3_Alignment_Score', 'DM3_Time', 'DM3_View', 'DM3_Review_Time',
               'Compare_Time']

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Write headers and rows
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    for row_idx, row_data in enumerate(table_data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

    # format
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    wb.save(output_file)
    print(f"Spreadsheet exported to {output_file}")
    print(f"Total number of participants processed: {len(processed_participants)}")
